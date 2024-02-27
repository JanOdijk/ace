import sys
import openpyxl
from openpyxl import load_workbook

baseversion = "0"
subversion = "16"
version = baseversion +  "." + subversion



def read_errors(fullname, error_annotations,message):
  wb = load_workbook(fullname)
  wsns = wb.sheetnames
  if wsns != []:
     ws=wb[wsns[0]]
  else:
    print("Error reading {}. No worksheets found\n".format(fullname), file=sys.stderr)   
  rowctr=-1
  for row in ws.iter_rows():
    rowctr+=1
    #skip header
    if rowctr!=0:
       if not(row[0].value is None and row[1].value is None) and row[2]!=message:
          r3v=nonetostr(row[3].value) #filename
          r4v=nonetostr(row[4].value) #mother:cat
          r5v=nonetostr(row[5].value) #node:rel/poscat
          r6v=str(nonetostr(row[6].value)) #errno
          r7v=nonetostr(row[7].value) #arg1
          r9v=nonetostr(row[9].value) #arg2
          #r10v=nonetostr(row[10].value)# username
          r0v=nonetostr(row[0].value) #user1
          r1v=nonetostr(row[1].value) #user2
          r10v=nonetostr(row[10].value) #annotator
#          error_annotations[(r3v,r4v,r5v,r6v,r7v,r9v)][r10v]=(r0v,r1v)
          error_annotations[(r3v,r4v,r5v,r6v,r7v,r9v)]=(r0v,r1v,r10v)
          
def nonetostr(x):
  if x is None:
    result=""
  else:
    result = x
  return(result)    
